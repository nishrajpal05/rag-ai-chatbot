


import logging
from pathlib import Path
from typing import Optional
import PyPDF2
import docx
import openpyxl

logger = logging.getLogger(__name__)

class DocumentProcessor:
    """Process different document types and extract text"""
    
    def extract_text(self, file_path: str) -> Optional[str]:
        """Extract text from various file formats"""
        try:
            file_path = Path(file_path)
            
            if not file_path.exists():
                logger.error(f"File not found: {file_path}")
                return None
            
            suffix = file_path.suffix.lower()
            
            logger.info(f" Extracting text from {suffix} file: {file_path.name}")
            
            if suffix == '.pdf':
                return self._extract_from_pdf(file_path)
            elif suffix in ['.docx', '.doc']:
                return self._extract_from_docx(file_path)
            elif suffix in ['.xlsx', '.xls']:
                return self._extract_from_excel(file_path)
            elif suffix == '.txt':
                return self._extract_from_txt(file_path)
            else:
                logger.warning(f"Unsupported file type: {suffix}")
                return None
                
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {str(e)}")
            return None
    
    def _extract_from_pdf(self, file_path: Path) -> str:
        """Extract text from PDF"""
        try:
            text = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                num_pages = len(pdf_reader.pages)
                
                logger.info(f" PDF has {num_pages} pages")
                
                for page_num in range(num_pages):
                    page = pdf_reader.pages[page_num]
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            
            logger.info(f" Extracted {len(text)} characters from PDF")
            return text.strip()
            
        except Exception as e:
            logger.error(f"Error reading PDF: {str(e)}")
            return ""
    
    def _extract_from_docx(self, file_path: Path) -> str:
        """Extract text from DOCX"""
        try:
            doc = docx.Document(file_path)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip()])
            
            logger.info(f" Extracted {len(text)} characters from DOCX")
            return text.strip()
            
        except Exception as e:
            logger.error(f"Error reading DOCX: {str(e)}")
            return ""
    
    def _extract_from_excel(self, file_path: Path) -> str:
        """Extract text from Excel"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"Sheet: {sheet_name}\n")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text_parts.append(row_text)
            
            text = "\n".join(text_parts)
            logger.info(f" Extracted {len(text)} characters from Excel")
            return text.strip()
            
        except Exception as e:
            logger.error(f"Error reading Excel: {str(e)}")
            return ""
    
    def _extract_from_txt(self, file_path: Path) -> str:
        """Extract text from TXT"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                text = file.read()
            
            logger.info(f" Extracted {len(text)} characters from TXT")
            return text.strip()
            
        except Exception as e:
            logger.error(f"Error reading TXT: {str(e)}")
            return ""

# Singleton instance
document_processor = DocumentProcessor()